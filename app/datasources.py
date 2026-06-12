"""
ITSM data layer.

Loads the four ServiceNow CSV exports (Incidents, Problems, Changes, Tasks)
into a local SQLite database and provides query helpers for cross-domain
correlation. The shared key across Incidents/Problems/Changes is `cmdb_ci`
(the affected configuration item), plus `assignment_group` and timestamps.

The database (data/itsm.db) is built lazily and rebuilt whenever a source CSV
is newer than the DB. Everything stays local — data/ is git-ignored.
"""

from __future__ import annotations

import csv
import re
import sqlite3
from datetime import datetime
from pathlib import Path

csv.field_size_limit(10_000_000)

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "itsm.db"

# Filename token (whitespace-delimited) -> table name.
SOURCES = {"INC": "incidents", "PRB": "problems", "CR": "changes", "TSK": "tasks"}

# Per-table: which raw column to normalise into which ISO datetime helper column.
DATE_COLUMNS = {
    "incidents": {"opened_at": "opened_iso", "resolved_at": "resolved_iso", "closed_at": "closed_iso"},
    "changes": {"sys_created_on": "created_iso", "expected_start": "start_iso", "closed_at": "closed_iso"},
    "problems": {"sys_created_on": "created_iso", "closed_at": "closed_iso"},
    "tasks": {"sys_created_on": "created_iso", "closed_at": "closed_iso"},
}

_DATE_FORMATS = (
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
    "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y",
    "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
)


def parse_dt(value: str) -> datetime | None:
    v = (value or "").strip()
    if not v or v.upper() == "UNKNOWN":
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(v, fmt)
        except ValueError:
            continue
    return None


def _iso(value: str) -> str | None:
    dt = parse_dt(value)
    return dt.isoformat(sep=" ") if dt else None


def _sanitize(col: str) -> str:
    """Make a CSV header safe as a SQLite column name."""
    c = re.sub(r"[^0-9a-zA-Z]+", "_", col.strip()).strip("_").lower()
    return c or "col"


def _find_csv(token: str) -> Path | None:
    for p in DATA_DIR.glob("*.csv"):
        if token in p.stem.replace("_", " ").split():
            return p
    return None


# --- Building the database --------------------------------------------------

def _needs_rebuild() -> bool:
    if not DB_PATH.exists():
        return True
    db_mtime = DB_PATH.stat().st_mtime
    for token in SOURCES:
        p = _find_csv(token)
        if p and p.stat().st_mtime > db_mtime:
            return True
    return False


def _load_table(conn: sqlite3.Connection, table: str, header, rows, date_map: dict) -> int:
    """Load a dataset (columns + row iterator from a connector) into `table`."""
    header = list(header or [])
    if not header:
        return 0
    cols = [_sanitize(h) for h in header]
    # De-duplicate any colliding sanitized names.
    seen: dict[str, int] = {}
    for i, c in enumerate(cols):
        if c in seen:
            seen[c] += 1
            cols[i] = f"{c}_{seen[c]}"
        else:
            seen[c] = 0

    extra_cols = list(date_map.values())
    all_cols = cols + extra_cols

    conn.execute(f"DROP TABLE IF EXISTS {table}")
    col_defs = ", ".join(f'"{c}" TEXT' for c in all_cols)
    conn.execute(f"CREATE TABLE {table} ({col_defs})")

    placeholders = ", ".join("?" for _ in all_cols)
    insert_sql = f"INSERT INTO {table} VALUES ({placeholders})"
    col_index = {c: i for i, c in enumerate(cols)}

    batch, total = [], 0
    for row in rows:
        row = (list(row) + [""] * len(cols))[: len(cols)]  # pad/trim to width
        extras = []
        for raw_col, iso_col in date_map.items():
            idx = col_index.get(_sanitize(raw_col))
            extras.append(_iso(row[idx]) if idx is not None else None)
        batch.append(row + extras)
        if len(batch) >= 5000:
            conn.executemany(insert_sql, batch)
            total += len(batch)
            batch.clear()
    if batch:
        conn.executemany(insert_sql, batch)
        total += len(batch)
    return total


def _create_indexes(conn: sqlite3.Connection) -> None:
    idx = [
        ("incidents", "cmdb_ci"), ("incidents", "assignment_group"), ("incidents", "opened_iso"),
        ("incidents", "category"), ("incidents", "priority"),
        ("changes", "cmdb_ci"), ("changes", "created_iso"),
        ("problems", "cmdb_ci"),
    ]
    for table, col in idx:
        try:
            conn.execute(f'CREATE INDEX IF NOT EXISTS ix_{table}_{col} ON {table}("{col}")')
        except sqlite3.OperationalError:
            pass  # column may not exist in a given export


def _build_fts(conn: sqlite3.Connection) -> bool:
    """Full-text index over incident text for similar-incident search."""
    try:
        cols = _table_columns(conn, "incidents")
        if "short_description" not in cols:
            return False  # nothing meaningful to index
        text_cols = [c for c in ("number", "short_description", "close_notes") if c in cols]
        joined = ", ".join(text_cols)
        conn.execute("DROP TABLE IF EXISTS incident_fts")
        conn.execute(
            f"CREATE VIRTUAL TABLE incident_fts USING fts5("
            f"{joined}, content='incidents', content_rowid='rowid')")
        conn.execute(
            f"INSERT INTO incident_fts(rowid, {joined}) SELECT rowid, {joined} FROM incidents")
        return True
    except sqlite3.OperationalError:
        return False  # FTS5 not compiled in / schema mismatch — callers fall back to LIKE


def build(force: bool = False) -> dict:
    """Build the SQLite DB from the active connector. Returns row counts per table."""
    from app.connectors import get_connector  # lazy import (avoids any import cycle)

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not force and not _needs_rebuild():
        return stats()
    connector = get_connector()
    counts: dict = {}
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        for table in SOURCES.values():  # incidents, problems, changes, tasks
            header, rows = connector.fetch(table)
            counts[table] = _load_table(conn, table, header, rows, DATE_COLUMNS.get(table, {}))
        _create_indexes(conn)
        counts["fts"] = _build_fts(conn)
        counts["source"] = connector.source_name
        conn.commit()
    finally:
        conn.close()
    return counts


def data_source() -> dict:
    """Active data source description (for /api/itsm/source and stats)."""
    from app.connectors import source_status

    return source_status()


_ensured = False


def ensure_loaded() -> None:
    global _ensured
    if not _ensured or _needs_rebuild():
        build()
        _ensured = True


def get_connection(db_path: str | Path | None = None) -> sqlite3.Connection:
    """Open the base DB (db_path=None, lazily built) or a specific DB (e.g. a
    per-session uploaded dataset)."""
    if db_path is None:
        ensure_loaded()
        db_path = DB_PATH
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


# --- Query helpers ----------------------------------------------------------

def _table_columns(conn: sqlite3.Connection, table: str) -> set[str]:
    return {r[1] for r in conn.execute(f"PRAGMA table_info({table})")}


def _present(conn: sqlite3.Connection, table: str, cols: list[str]) -> list[str]:
    """Subset of `cols` that actually exist in `table` (handles varied uploads)."""
    have = _table_columns(conn, table)
    return [c for c in cols if c in have]


def _count(conn: sqlite3.Connection, table: str, col: str, val) -> int:
    try:
        return conn.execute(f'SELECT COUNT(*) FROM {table} WHERE "{col}" = ?', (val,)).fetchone()[0]
    except sqlite3.OperationalError:
        return 0


def stats(db_path: str | Path | None = None) -> dict:
    path = Path(db_path) if db_path else DB_PATH
    if not path.exists():
        return {}
    conn = sqlite3.connect(path)
    try:
        out = {}
        for table in SOURCES.values():
            try:
                out[table] = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            except sqlite3.OperationalError:
                out[table] = 0
        return out
    finally:
        conn.close()


def get_incident(number: str, db_path: str | Path | None = None) -> dict | None:
    conn = get_connection(db_path)
    try:
        row = conn.execute("SELECT * FROM incidents WHERE number = ?", (number.strip(),)).fetchone()
        return dict(row) if row else None
    finally:
        conn.close()


def correlate_ci(ci: str, limit: int = 50, db_path: str | Path | None = None) -> dict:
    """All incidents / problems / changes touching a configuration item."""
    conn = get_connection(db_path)

    def _ci_rows(table, wanted, order_pref):
        cols = _present(conn, table, wanted)
        if not cols or "cmdb_ci" not in _table_columns(conn, table):
            return []
        order = order_pref if order_pref in cols else cols[0]
        return [dict(r) for r in conn.execute(
            f"SELECT {', '.join(cols)} FROM {table} WHERE cmdb_ci = ? "
            f"ORDER BY {order} DESC LIMIT ?", (ci, limit))]

    try:
        out: dict = {"cmdb_ci": ci}
        out["incidents"] = _ci_rows("incidents",
            ["number", "priority", "short_description", "category", "assignment_group",
             "opened_iso", "closed_iso", "close_code"], "opened_iso")
        out["incident_count"] = _count(conn, "incidents", "cmdb_ci", ci)
        out["problems"] = _ci_rows("problems",
            ["number", "short_description", "state", "resolution_code", "related_incidents",
             "created_iso"], "created_iso")
        out["changes"] = _ci_rows("changes",
            ["number", "type", "approval", "impact", "created_iso", "start_iso", "closed_iso"],
            "created_iso")
        out["change_count"] = _count(conn, "changes", "cmdb_ci", ci)
        return out
    finally:
        conn.close()


def hotspots(top: int = 10, db_path: str | Path | None = None) -> dict:
    """Aggregate views for the dashboard."""
    conn = get_connection(db_path)
    try:
        out: dict = {}
        cols = _table_columns(conn, "incidents")
        out["top_cis"] = [dict(r) for r in conn.execute(
            "SELECT cmdb_ci AS ci, COUNT(*) AS incidents FROM incidents "
            "WHERE cmdb_ci IS NOT NULL AND cmdb_ci != '' "
            "GROUP BY cmdb_ci ORDER BY incidents DESC LIMIT ?", (top,))] if "cmdb_ci" in cols else []
        out["top_groups"] = [dict(r) for r in conn.execute(
            "SELECT assignment_group AS team, COUNT(*) AS incidents FROM incidents "
            "WHERE assignment_group IS NOT NULL AND assignment_group != '' "
            "GROUP BY assignment_group ORDER BY incidents DESC LIMIT ?", (top,))] if "assignment_group" in cols else []
        out["top_categories"] = [dict(r) for r in conn.execute(
            "SELECT category, COUNT(*) AS incidents FROM incidents "
            "WHERE category != '' GROUP BY category ORDER BY incidents DESC LIMIT ?", (top,))] if "category" in cols else []
        out["by_month"] = [dict(r) for r in conn.execute(
            "SELECT substr(opened_iso,1,7) AS month, COUNT(*) AS incidents FROM incidents "
            "WHERE opened_iso IS NOT NULL GROUP BY month ORDER BY month")] if "opened_iso" in cols else []
        if "made_sla" in cols:
            total = conn.execute("SELECT COUNT(*) FROM incidents WHERE made_sla IN ('true','false')").fetchone()[0]
            breached = conn.execute("SELECT COUNT(*) FROM incidents WHERE made_sla = 'false'").fetchone()[0]
            out["sla"] = {"total": total, "breached": breached,
                          "breach_pct": round(100 * breached / total, 1) if total else 0}
        if "reopen_count" in cols:
            out["reopened"] = conn.execute(
                "SELECT COUNT(*) FROM incidents WHERE reopen_count NOT IN ('', '0')").fetchone()[0]
        if "major_incident_state" in cols:
            out["major_incidents"] = conn.execute(
                "SELECT COUNT(*) FROM incidents WHERE major_incident_state NOT IN ('', 'No')").fetchone()[0]
        return out
    finally:
        conn.close()


def list_cis(limit: int = 1000, db_path: str | Path | None = None) -> list[dict]:
    """Distinct configuration items, ordered by incident volume (for dropdowns)."""
    conn = get_connection(db_path)
    try:
        if "cmdb_ci" not in _table_columns(conn, "incidents"):
            return []
        return [dict(r) for r in conn.execute(
            "SELECT cmdb_ci AS ci, COUNT(*) AS incidents FROM incidents "
            "WHERE cmdb_ci IS NOT NULL AND cmdb_ci != '' "
            "GROUP BY cmdb_ci ORDER BY incidents DESC LIMIT ?", (limit,))]
    finally:
        conn.close()


# Incident columns that make sense to break down by (whitelist — used in SQL).
BREAKDOWN_FIELDS = {
    "cmdb_ci": "Configuration Item",
    "assignment_group": "Assignment Group",
    "category": "Category",
    "subcategory": "Subcategory",
    "priority": "Priority",
    "incident_state": "State",
    "contact_type": "Contact Type",
}


def breakdown(by: str = "category", top: int = 12, db_path: str | Path | None = None) -> dict:
    """Incident counts grouped by a whitelisted dimension."""
    conn = get_connection(db_path)
    try:
        cols = _table_columns(conn, "incidents")
        if by not in BREAKDOWN_FIELDS or by not in cols:
            by = "category" if "category" in cols else "cmdb_ci"
        rows = conn.execute(
            f'SELECT "{by}" AS label, COUNT(*) AS count FROM incidents '
            f'WHERE "{by}" IS NOT NULL AND "{by}" != \'\' '
            f'GROUP BY "{by}" ORDER BY count DESC LIMIT ?', (top,)).fetchall()
        return {"by": by, "label": BREAKDOWN_FIELDS.get(by, by), "rows": [dict(r) for r in rows]}
    finally:
        conn.close()


def change_incident_correlation(window_hours: int = 72, top: int = 15,
                                db_path: str | Path | None = None) -> list[dict]:
    """Find changes followed by incidents on the same CI within `window_hours`.

    A strong signal for change-induced incidents.
    """
    conn = get_connection(db_path)
    try:
        rows = conn.execute(
            "SELECT c.number AS change, c.type, c.cmdb_ci AS ci, c.created_iso AS change_time, "
            "       COUNT(i.number) AS incidents_after "
            "FROM changes c JOIN incidents i "
            "  ON i.cmdb_ci = c.cmdb_ci "
            " AND i.opened_iso IS NOT NULL AND c.created_iso IS NOT NULL "
            " AND i.opened_iso >= c.created_iso "
            " AND julianday(i.opened_iso) - julianday(c.created_iso) <= ?/24.0 "
            "WHERE c.cmdb_ci IS NOT NULL AND c.cmdb_ci != '' "
            "GROUP BY c.number HAVING incidents_after > 0 "
            "ORDER BY incidents_after DESC LIMIT ?", (window_hours, top)).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def similar_incidents(text: str, k: int = 8, db_path: str | Path | None = None) -> list[dict]:
    """Find past incidents most similar to `text` (FTS5, with LIKE fallback)."""
    conn = get_connection(db_path)
    try:
        if "short_description" not in _table_columns(conn, "incidents"):
            return []
        sel = _present(conn, "incidents", ["number", "short_description", "category",
              "close_code", "close_notes", "cmdb_ci", "assignment_group"])
        has_fts = bool(conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='incident_fts'").fetchone())
        terms = re.findall(r"[A-Za-z0-9]{3,}", text.lower())
        if has_fts and terms:
            query = " OR ".join(terms[:20])
            try:
                sel_sql = ", ".join(f"i.{c}" for c in sel)
                rows = conn.execute(
                    f"SELECT {sel_sql} FROM incident_fts f JOIN incidents i ON i.rowid = f.rowid "
                    f"WHERE incident_fts MATCH ? ORDER BY bm25(incident_fts) LIMIT ?", (query, k)).fetchall()
                if rows:
                    return [dict(r) for r in rows]
            except sqlite3.OperationalError:
                pass
        # Fallback: LIKE on the longest terms.
        like_terms = sorted(set(terms), key=len, reverse=True)[:3] or [text[:20]]
        clause = " OR ".join("short_description LIKE ?" for _ in like_terms)
        params = [f"%{t}%" for t in like_terms] + [k]
        rows = conn.execute(
            f"SELECT {', '.join(sel)} FROM incidents WHERE {clause} LIMIT ?", params).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


# --- Per-session uploaded datasets ------------------------------------------

SESSIONS_DIR = DATA_DIR / "sessions"


def _safe_sid(sid: str | None) -> str:
    return re.sub(r"[^A-Za-z0-9_-]", "", sid or "")[:64]


def session_db_path(sid: str) -> Path | None:
    s = _safe_sid(sid)
    return SESSIONS_DIR / f"{s}.db" if s else None


def session_exists(sid: str) -> bool:
    p = session_db_path(sid)
    return bool(p and p.exists())


def _read_tabular(filename: str, data: bytes):
    """Return (header, rows) from an uploaded CSV/TSV or XLSX file."""
    import io

    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext in ("xlsx", "xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = ["" if v is None else str(v) for v in next(it, [])]

        def rows():
            try:
                for r in it:
                    yield ["" if v is None else str(v) for v in r]
            finally:
                wb.close()

        return header, rows()

    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text), delimiter="\t" if ext == "tsv" else ",")
    header = next(reader, [])
    return header, reader


def _detect_dataset(filename: str, header: list[str]) -> str | None:
    """Infer which dataset a file is, from its filename token or column signature."""
    tokens = set(re.split(r"[^A-Z0-9]+", filename.rsplit(".", 1)[0].upper()))
    if {"INC", "INCIDENT", "INCIDENTS"} & tokens:
        return "incidents"
    if {"PRB", "PROBLEM", "PROBLEMS"} & tokens:
        return "problems"
    if {"CR", "CHG", "CHANGE", "CHANGES"} & tokens:
        return "changes"
    if {"TSK", "SCTASK", "TASK", "TASKS"} & tokens:
        return "tasks"
    cols = {_sanitize(h) for h in header}
    if {"resolution_code", "related_incidents"} & cols:
        return "problems"
    if {"backout_plan", "approval"} & cols:
        return "changes"
    if {"request_item", "cat_item"} & cols:
        return "tasks"
    if {"incident_state", "severity", "close_code"} & cols:
        return "incidents"
    return None


def build_session_db(sid: str, files: list) -> dict:
    """Build an isolated SQLite DB for a session from uploaded files.

    `files` is a list of (filename, bytes); each file's dataset is auto-detected.
    Returns the detected mapping + per-table row counts.
    """
    s = _safe_sid(sid)
    if not s:
        raise ValueError("A valid session id is required.")
    SESSIONS_DIR.mkdir(parents=True, exist_ok=True)
    path = SESSIONS_DIR / f"{s}.db"
    if path.exists():
        path.unlink()
    detected: dict = {}
    counts: dict = {}
    conn = sqlite3.connect(path)
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        for filename, data in files:
            header, rows = _read_tabular(filename, data)
            dataset = _detect_dataset(filename, header) if header else None
            detected[filename] = dataset or "unrecognized"
            if dataset:
                counts[dataset] = _load_table(conn, dataset, header, rows,
                                              DATE_COLUMNS.get(dataset, {}))
        _create_indexes(conn)
        counts["fts"] = _build_fts(conn)
        conn.commit()
    finally:
        conn.close()
    return {"detected": detected, "counts": counts}
